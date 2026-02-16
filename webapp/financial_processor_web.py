"""
財務諸表プロセッサ（Web版）
アップロードされた決算書PDFから財務情報を抽出してExcelに書き込む
"""
from io import BytesIO
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from utils.claude_client import ClaudeClient


class FinancialProcessorWeb:
    """財務諸表プロセッサ（Web版）"""

    def __init__(self):
        self.claude_client = ClaudeClient()

    def process_uploaded_files(
        self,
        period1_pdf,
        period2_pdf,
        period3_pdf,
        account_details_pdf,
        company_name: str
    ) -> BytesIO:
        """
        アップロードされた決算書PDFから財務概要Excelを生成

        Args:
            period1_pdf: Period 1（最古）のPDF
            period2_pdf: Period 2（中間）のPDF
            period3_pdf: Period 3（最新）のPDF
            account_details_pdf: 科目明細のPDF（最新期）
            company_name: 会社名

        Returns:
            BytesIO: Excelファイルのバイトストリーム
        """
        # 各期の財務データ抽出
        period1_data = self._extract_financial_data(period1_pdf, 1) if period1_pdf else None
        period2_data = self._extract_financial_data(period2_pdf, 2) if period2_pdf else None
        period3_data = self._extract_financial_data(period3_pdf, 3) if period3_pdf else None

        # 科目明細抽出（最新期のみ）
        account_details = self._extract_account_details(account_details_pdf) if account_details_pdf else None

        # Excelワークブック作成
        wb = self._create_workbook()

        # PLシートへの書き込み
        self._write_pl_sheet(wb, period1_data, period2_data, period3_data)

        # BSシートへの書き込み
        self._write_bs_sheet(wb, period1_data, period2_data, period3_data)

        # 科目シートへの書き込み
        if account_details:
            self._write_account_sheet(wb, account_details)

        # BytesIOに保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def _extract_financial_data(self, pdf_file, period_num: int) -> Dict[str, Any]:
        """
        決算書PDFから財務データを抽出

        Args:
            pdf_file: StreamlitのUploadedFileオブジェクト
            period_num: 期番号（1, 2, 3）

        Returns:
            財務データの辞書
        """
        pdf_bytes = pdf_file.read()
        prompt = self._get_financial_extraction_prompt()

        result = self.claude_client.analyze_pdf_bytes(pdf_bytes, prompt)

        if isinstance(result, dict) and not result.get("error"):
            result["period_num"] = period_num
            return result
        else:
            return {"error": f"Period {period_num}のデータ抽出に失敗しました", "period_num": period_num}

    def _extract_account_details(self, pdf_file) -> List[Dict[str, Any]]:
        """
        科目明細PDFから詳細データを抽出

        Args:
            pdf_file: StreamlitのUploadedFileオブジェクト

        Returns:
            科目明細のリスト
        """
        pdf_bytes = pdf_file.read()
        prompt = self._get_account_details_prompt()

        result = self.claude_client.analyze_pdf_bytes(pdf_bytes, prompt)

        if isinstance(result, list):
            return result
        elif isinstance(result, dict) and not result.get("error"):
            # 単一オブジェクトの場合はリストに変換
            return [result]
        else:
            return []

    def _get_financial_extraction_prompt(self) -> str:
        """決算書解析用プロンプト"""
        return """
この決算書PDFから以下の情報を抽出してJSONで返してください。

## 抽出する情報

### 1. 決算期情報
- 期数（例: 第25期）
- 決算日（例: 2024年9月30日）

### 2. 損益計算書（PL）
以下の科目を抽出してください。金額が記載されていない場合は0を設定してください。

- 売上高
- 売上総利益
- 売上原価
- 売上総利益率(%)
- 営業利益
- 営業利益率(%)
- 経常利益
- 当期純利益

### 3. 貸借対照表（BS）
以下の科目を抽出してください。金額が記載されていない場合は0を設定してください。

- 流動資産
- 固定資産
- 資産合計
- 流動負債
- 固定負債
- 負債合計
- 純資産
- 負債・純資産合計

### JSONフォーマット:
{
  "fiscal_period": "第25期",
  "fiscal_year_end": "2024-09-30",
  "pl": {
    "売上高": 111899685,
    "売上総利益": 11940879,
    "売上原価": 99958806,
    "売上総利益率": 10.7,
    "営業利益": -3947829,
    "営業利益率": -3.5,
    "経常利益": -4195663,
    "当期純利益": -3307208
  },
  "bs": {
    "流動資産": 11539877,
    "固定資産": 2337566,
    "資産合計": 13877443,
    "流動負債": 18577175,
    "固定負債": 7500000,
    "負債合計": 26077175,
    "純資産": -12199732,
    "負債・純資産合計": 13877443
  }
}

必ずJSON形式で返してください。コードブロックは使わず、JSONのみを返してください。
"""

    def _get_account_details_prompt(self) -> str:
        """科目明細解析用プロンプト"""
        return """
この科目明細PDFから全ての科目情報を抽出してJSON配列で返してください。

## 抽出する情報
各科目について以下を抽出:
- 科目名（例: 現金、普通預金）
- 詳細（例: 東京三菱UFJ銀行 本店営業部、備考情報など）
- 金額

### JSONフォーマット（配列で返す）:
[
  {
    "科目名": "現金",
    "詳細": "現金",
    "金額": 10473
  },
  {
    "科目名": "普通預金",
    "詳細": "東京三菱UFJ銀行 本店営業部",
    "金額": 1402557
  },
  {
    "科目名": "普通預金",
    "詳細": "三井住友銀行 渋谷支店",
    "金額": 853
  }
]

必ずJSON配列形式で返してください。コードブロックは使わず、JSONのみを返してください。
"""

    def _create_workbook(self) -> Workbook:
        """
        新規Workbookを作成（PLシート、BSシート、科目シート）

        Returns:
            Workbookオブジェクト
        """
        wb = openpyxl.Workbook()

        # 最初のシートを削除
        wb.remove(wb.active)

        # PLシート作成
        ws_pl = wb.create_sheet("PL")
        self._setup_pl_sheet(ws_pl)

        # BSシート作成
        ws_bs = wb.create_sheet("BS")
        self._setup_bs_sheet(ws_bs)

        # 科目シート作成
        ws_account = wb.create_sheet("科目")
        self._setup_account_sheet(ws_account)

        return wb

    def _setup_pl_sheet(self, ws):
        """PLシートのヘッダー設定"""
        # ヘッダー行（1行目）
        ws.cell(1, 3, "決算科目")

        # 期間行（2行目）
        ws.cell(2, 2, "損益計算書")
        ws.cell(2, 5, "第○期\nYY.MM.DD")
        ws.cell(2, 6, "第○期\nYY.MM.DD")
        ws.cell(2, 7, "第○期\nYY.MM.DD")

        # 科目ヘッダー（3行目以降）
        pl_items = [
            ("売上高", "売上原価"),
            ("売上総利益", ""),
            ("販管費", ""),
            ("営業利益", ""),
            ("営業外収益", ""),
            ("営業外費用", ""),
            ("経常利益", ""),
            ("特別利益", ""),
            ("特別損失", ""),
            ("税引前当期純利益", ""),
            ("法人税等", ""),
            ("当期純利益", ""),
        ]

        row = 3
        for major, minor in pl_items:
            ws.cell(row, 2, major)
            if minor:
                ws.cell(row, 3, minor)
            row += 1

    def _setup_bs_sheet(self, ws):
        """BSシートのヘッダー設定"""
        # ヘッダー行（1行目）
        ws.cell(1, 3, "決算科目")

        # 期間行（2行目）
        ws.cell(2, 2, "貸借対照表")
        ws.cell(2, 5, "第○期\nYY.MM.DD")
        ws.cell(2, 6, "第○期\nYY.MM.DD")
        ws.cell(2, 7, "第○期\nYY.MM.DD")

        # 科目ヘッダー（3行目以降）
        bs_items = [
            ("流動資産", "現金"),
            ("", "預金"),
            ("", "売掛金"),
            ("", "棚卸資産"),
            ("", "その他流動資産"),
            ("固定資産", "有形固定資産"),
            ("", "無形固定資産"),
            ("", "投資その他の資産"),
            ("資産合計", ""),
            ("流動負債", "買掛金"),
            ("", "短期借入金"),
            ("", "未払金"),
            ("", "その他流動負債"),
            ("固定負債", "長期借入金"),
            ("", "その他固定負債"),
            ("負債合計", ""),
            ("純資産", "資本金"),
            ("", "利益剰余金"),
            ("負債・純資産合計", ""),
        ]

        row = 3
        for major, minor in bs_items:
            ws.cell(row, 2, major)
            if minor:
                ws.cell(row, 3, minor)
            row += 1

    def _setup_account_sheet(self, ws):
        """科目シートのヘッダー設定"""
        # ヘッダー行
        ws.cell(2, 1, "No.")
        ws.cell(2, 2, "科目")
        ws.cell(2, 3, "詳細")
        ws.cell(2, 4, "金額")
        ws.cell(2, 5, "備考1")
        ws.cell(2, 6, "備考2")

    def _write_pl_sheet(self, wb: Workbook, period1_data, period2_data, period3_data):
        """PLシートへのデータ書き込み"""
        ws = wb["PL"]

        periods = [period1_data, period2_data, period3_data]

        for idx, period_data in enumerate(periods, start=5):  # 列E, F, G
            if not period_data or period_data.get("error"):
                continue

            col = idx  # 5, 6, 7

            # 期間情報を2行目に書き込み
            fiscal_period = period_data.get("fiscal_period", "")
            fiscal_year_end = period_data.get("fiscal_year_end", "")
            ws.cell(2, col, f"{fiscal_period}\n{fiscal_year_end}")

            # PLデータ
            pl = period_data.get("pl", {})

            # データマッピング（行番号: キー名）
            pl_mapping = {
                3: "売上高",
                4: "売上原価",
                5: "売上総利益",
                6: "販管費",
                7: "営業利益",
                11: "経常利益",
                14: "当期純利益",
            }

            for row, key in pl_mapping.items():
                value = pl.get(key, 0)
                ws.cell(row, col, value)

    def _write_bs_sheet(self, wb: Workbook, period1_data, period2_data, period3_data):
        """BSシートへのデータ書き込み"""
        ws = wb["BS"]

        periods = [period1_data, period2_data, period3_data]

        for idx, period_data in enumerate(periods, start=5):  # 列E, F, G
            if not period_data or period_data.get("error"):
                continue

            col = idx  # 5, 6, 7

            # 期間情報を2行目に書き込み
            fiscal_period = period_data.get("fiscal_period", "")
            fiscal_year_end = period_data.get("fiscal_year_end", "")
            ws.cell(2, col, f"{fiscal_period}\n{fiscal_year_end}")

            # BSデータ
            bs = period_data.get("bs", {})

            # データマッピング（行番号: キー名）
            bs_mapping = {
                3: "流動資産",
                10: "固定資産",
                18: "資産合計",
                19: "流動負債",
                24: "固定負債",
                26: "負債合計",
                27: "純資産",
                30: "負債・純資産合計",
            }

            for row, key in bs_mapping.items():
                value = bs.get(key, 0)
                ws.cell(row, col, value)

    def _write_account_sheet(self, wb: Workbook, account_details: List[Dict]):
        """科目シートへのデータ書き込み"""
        ws = wb["科目"]

        row = 3  # データは3行目から

        for idx, account in enumerate(account_details, start=1):
            ws.cell(row, 1, idx)  # No.
            ws.cell(row, 2, account.get("科目名", ""))
            ws.cell(row, 3, account.get("詳細", ""))
            ws.cell(row, 4, account.get("金額", 0))
            row += 1
