"""
IM（Information Memorandum）下書き自動生成プロセッサ（Web版）
複数の資料から構造化されたIM下書きをExcel形式で出力
"""
from io import BytesIO
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from utils.claude_client import ClaudeClient


class IMGeneratorWeb:
    """IM下書き自動生成プロセッサ（Web版）"""

    def __init__(self):
        self.claude_client = ClaudeClient()

    def process_uploaded_files(
        self,
        uploaded_files: List,
        company_name: str
    ) -> BytesIO:
        """
        アップロードされた資料からIM下書きExcelを生成

        Args:
            uploaded_files: Streamlit UploadedFileのリスト（決算書、会社案内、組織図など）
            company_name: 会社名

        Returns:
            BytesIO: Excelファイルのバイトストリーム
        """
        # 各ファイルを解析
        analyzed_data = []

        for uploaded_file in uploaded_files:
            file_bytes = uploaded_file.read()
            filename = uploaded_file.name

            # ファイル種別を判定して解析
            if filename.lower().endswith('.pdf'):
                file_info = self.claude_client.analyze_pdf_bytes(
                    file_bytes, self._get_im_analysis_prompt()
                )
            else:
                # 画像ファイル
                file_info = self.claude_client.analyze_image_bytes(
                    file_bytes, filename, self._get_im_analysis_prompt()
                )

            if isinstance(file_info, dict) and not file_info.get("error"):
                file_info["filename"] = filename
                analyzed_data.append(file_info)

        # 情報を統合・構造化
        structured_data = self._structure_im_data(analyzed_data, company_name)

        # Excelワークブック作成
        wb = self._create_im_workbook(structured_data)

        # BytesIOに保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def _get_im_analysis_prompt(self) -> str:
        """IM資料解析用プロンプト"""
        return """
この資料を解析して、以下の情報をJSONで返してください。

## 資料種別の判定
まず、この資料が何の資料かを判定してください：
- "financial": 決算書、財務諸表（PL、BS、販管費、原価内訳など）
- "company_overview": 会社案内、パンフレット、会社概要
- "organization": 組織図、従業員リスト
- "assets": 資産リスト（車両台帳、不動産、設備など）
- "debt": 借入金明細、負債リスト
- "clients": 取引先リスト、顧客リスト
- "other": その他

## 抽出する情報
資料種別に応じて、以下の情報を抽出してください：

### financial（財務諸表）の場合:
- 期数（例: 第25期）
- 決算日
- 売上高、営業利益、経常利益、当期純利益
- 資産合計、負債合計、純資産

### company_overview（会社概要）の場合:
- 会社名
- 所在地
- 代表者名
- 設立年月日
- 資本金
- 事業内容（簡潔に）
- 従業員数
- 主要取引先
- 会社の強み・特徴

### organization（組織・従業員）の場合:
- 従業員数
- 部門構成
- 役員・幹部の氏名と役職
- 平均年齢
- 平均勤続年数

### assets（資産）の場合:
- 資産種別（車両、不動産、設備など）
- 資産の概要
- 評価額（あれば）

### debt（負債）の場合:
- 借入先
- 借入残高
- 返済期限

### clients（取引先）の場合:
- 主要取引先の名称
- 取引内容
- 取引金額（あれば）

### other（その他）の場合:
- 内容の要約

JSONフォーマット:
{
  "document_type": "financial/company_overview/organization/assets/debt/clients/other",
  "content": {
    ... (資料種別に応じた情報)
  },
  "summary": "この資料の要約（1-2行）"
}

重要:
- 資料種別は必ず判定してください
- 情報がない項目は空文字列""を設定
- 必ずJSON形式で返してください（コードブロック不要）
"""

    def _structure_im_data(
        self,
        analyzed_data: List[Dict[str, Any]],
        company_name: str
    ) -> Dict[str, Any]:
        """
        解析データを構造化
        """
        structured = {
            "company_name": company_name,
            "company_overview": {},
            "financial": [],
            "organization": {},
            "assets": [],
            "debt": [],
            "clients": [],
            "other": []
        }

        for data in analyzed_data:
            doc_type = data.get("document_type", "other")
            content = data.get("content", {})

            if doc_type == "financial":
                structured["financial"].append(content)
            elif doc_type == "company_overview":
                # 会社概要は最初のもので上書き
                if not structured["company_overview"]:
                    structured["company_overview"] = content
            elif doc_type == "organization":
                structured["organization"] = content
            elif doc_type == "assets":
                structured["assets"].append(content)
            elif doc_type == "debt":
                structured["debt"].append(content)
            elif doc_type == "clients":
                structured["clients"].append(content)
            else:
                structured["other"].append(data)

        return structured

    def _create_im_workbook(self, structured_data: Dict[str, Any]) -> Workbook:
        """
        IM下書きExcelワークブックを作成
        """
        wb = openpyxl.Workbook()

        # デフォルトシートを削除
        wb.remove(wb.active)

        # シート1: 会社概要
        self._create_overview_sheet(wb, structured_data)

        # シート2: 財務サマリー
        self._create_financial_sheet(wb, structured_data)

        # シート3: 組織・従業員
        self._create_organization_sheet(wb, structured_data)

        # シート4: 資産・負債
        self._create_assets_sheet(wb, structured_data)

        # シート5: その他情報
        self._create_other_sheet(wb, structured_data)

        return wb

    def _create_overview_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """会社概要シート作成"""
        ws = wb.create_sheet("会社概要")

        # タイトル
        ws['A1'] = f"{data['company_name']} 会社概要"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')

        overview = data.get("company_overview", {})

        # データ項目
        items = [
            ("会社名", overview.get("会社名", data['company_name'])),
            ("所在地", overview.get("所在地", "")),
            ("代表者", overview.get("代表者名", "")),
            ("設立年月日", overview.get("設立年月日", "")),
            ("資本金", overview.get("資本金", "")),
            ("従業員数", overview.get("従業員数", "")),
            ("事業内容", overview.get("事業内容", "")),
            ("主要取引先", overview.get("主要取引先", "")),
            ("会社の強み", overview.get("会社の強み・特徴", "")),
        ]

        row = 3
        for label, value in items:
            ws.cell(row, 1).value = label
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 2).value = value
            row += 1

        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60

    def _create_financial_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """財務サマリーシート作成"""
        ws = wb.create_sheet("財務サマリー")

        # タイトル
        ws['A1'] = f"{data['company_name']} 財務サマリー"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        # ヘッダー
        headers = ["科目", "第1期", "第2期", "第3期", "備考"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(3, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # 財務データ
        financial_data = data.get("financial", [])

        # PL項目
        pl_items = ["売上高", "営業利益", "経常利益", "当期純利益"]
        row = 4
        ws.cell(row, 1).value = "【損益計算書】"
        ws.cell(row, 1).font = Font(bold=True, italic=True)
        row += 1

        for item in pl_items:
            ws.cell(row, 1).value = item
            for idx, period in enumerate(financial_data[:3], start=2):
                value = period.get(item, 0)
                ws.cell(row, idx).value = value
            row += 1

        row += 1
        ws.cell(row, 1).value = "【貸借対照表】"
        ws.cell(row, 1).font = Font(bold=True, italic=True)
        row += 1

        # BS項目
        bs_items = ["資産合計", "負債合計", "純資産"]
        for item in bs_items:
            ws.cell(row, 1).value = item
            for idx, period in enumerate(financial_data[:3], start=2):
                value = period.get(item, 0)
                ws.cell(row, idx).value = value
            row += 1

        # 列幅調整
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 15

    def _create_organization_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """組織・従業員シート作成"""
        ws = wb.create_sheet("組織・従業員")

        ws['A1'] = f"{data['company_name']} 組織・従業員"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        org = data.get("organization", {})

        items = [
            ("従業員数", org.get("従業員数", "")),
            ("部門構成", org.get("部門構成", "")),
            ("役員・幹部", org.get("役員・幹部の氏名と役職", "")),
            ("平均年齢", org.get("平均年齢", "")),
            ("平均勤続年数", org.get("平均勤続年数", "")),
        ]

        row = 3
        for label, value in items:
            ws.cell(row, 1).value = label
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 2).value = value
            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60

    def _create_assets_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """資産・負債シート作成"""
        ws = wb.create_sheet("資産・負債")

        ws['A1'] = f"{data['company_name']} 資産・負債"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        row = 3

        # 資産セクション
        ws.cell(row, 1).value = "【資産】"
        ws.cell(row, 1).font = Font(bold=True, italic=True)
        row += 1

        assets = data.get("assets", [])
        for asset in assets:
            ws.cell(row, 1).value = asset.get("資産種別", "")
            ws.cell(row, 2).value = asset.get("資産の概要", "")
            ws.cell(row, 3).value = asset.get("評価額", "")
            row += 1

        row += 1

        # 負債セクション
        ws.cell(row, 1).value = "【負債】"
        ws.cell(row, 1).font = Font(bold=True, italic=True)
        row += 1

        debts = data.get("debt", [])
        for debt in debts:
            ws.cell(row, 1).value = debt.get("借入先", "")
            ws.cell(row, 2).value = debt.get("借入残高", "")
            ws.cell(row, 3).value = debt.get("返済期限", "")
            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15

    def _create_other_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """その他情報シート作成"""
        ws = wb.create_sheet("その他情報")

        ws['A1'] = f"{data['company_name']} その他情報"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        row = 3

        # 取引先
        clients = data.get("clients", [])
        if clients:
            ws.cell(row, 1).value = "【主要取引先】"
            ws.cell(row, 1).font = Font(bold=True, italic=True)
            row += 1

            for client in clients:
                ws.cell(row, 1).value = client.get("主要取引先の名称", "")
                ws.cell(row, 2).value = client.get("取引内容", "")
                row += 1

            row += 1

        # その他
        others = data.get("other", [])
        if others:
            ws.cell(row, 1).value = "【その他】"
            ws.cell(row, 1).font = Font(bold=True, italic=True)
            row += 1

            for other in others:
                ws.cell(row, 1).value = other.get("filename", "")
                ws.cell(row, 2).value = other.get("summary", "")
                row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
