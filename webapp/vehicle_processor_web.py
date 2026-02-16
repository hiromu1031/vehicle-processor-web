"""
車両台帳プロセッサ（Web版）
アップロードされたファイルから車両情報を抽出してExcelに書き込む
"""
from io import BytesIO
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook

from utils.claude_client import ClaudeClient


class VehicleProcessorWeb:
    """車両台帳プロセッサ（Web版）"""

    def __init__(self):
        self.claude_client = ClaudeClient()

    def process_uploaded_files(
        self, uploaded_files: List, company_name: str
    ) -> BytesIO:
        """
        アップロードされたファイルから車両台帳Excelを生成

        Args:
            uploaded_files: Streamlit UploadedFileのリスト
            company_name: 会社名

        Returns:
            BytesIO: Excelファイルのバイトストリーム
        """
        vehicles_data = []
        vehicle_no = 1

        # 各ファイルを処理
        for uploaded_file in uploaded_files:
            file_bytes = uploaded_file.read()
            filename = uploaded_file.name

            # 車両情報抽出
            if filename.lower().endswith('.pdf'):
                vehicle_info = self.claude_client.analyze_pdf_bytes(
                    file_bytes, self._get_vehicle_prompt()
                )
            else:
                vehicle_info = self.claude_client.analyze_image_bytes(
                    file_bytes, filename, self._get_vehicle_prompt()
                )

            # データ整形
            if isinstance(vehicle_info, list):
                for info in vehicle_info:
                    info["No"] = vehicle_no
                    vehicles_data.append(info)
                    vehicle_no += 1
            elif isinstance(vehicle_info, dict) and not vehicle_info.get("error"):
                vehicle_info["No"] = vehicle_no
                vehicles_data.append(vehicle_info)
                vehicle_no += 1

        # Excelワークブック作成
        wb = self._create_workbook(company_name)
        self._write_to_excel(wb, vehicles_data)

        # BytesIOに保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def _get_vehicle_prompt(self) -> str:
        """車検証解析用プロンプト"""
        return """
この車検証から以下の情報を抽出してJSONで返してください。
情報が読み取れない場合は空文字列""を設定してください。
複数台の車両情報がある場合は、すべて抽出して配列で返してください。

抽出する情報:
- 車両名 (例: ダイハツ ハイゼット)
- 車台番号
- 型式
- 自動車登録番号 (ナンバープレート)
- 新規登録年月
- 自動車種類 (例: 貨物、乗用)
- 車体の形状 (例: バン)
- 車名
- 最大積載量(kg)

JSONフォーマット(1台の場合):
{
  "車両名": "ダイハツ ハイゼット",
  "車台番号": "S331V-0046341",
  "型式": "EBD-S331V",
  "自動車登録番号": "群馬 480 ま 8601",
  "新規登録年月": "令和23年4月",
  "自動車種類": "貨物",
  "車体の形状": "バン",
  "車名": "ダイハツ",
  "最大積載量(kg)": "350(250)"
}

複数台の場合は配列で返してください:
[
  {...},
  {...}
]

必ずJSON形式で返してください。コードブロックは使わず、JSONのみを返してください。
"""

    def _create_workbook(self, company_name: str) -> Workbook:
        """
        新規Workbookを作成

        Args:
            company_name: 会社名

        Returns:
            Workbookオブジェクト
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "車両台帳"

        # ヘッダー行を作成(2行目)
        headers = [
            "No",
            "ステータス",
            "車両名",
            "車台番号",
            "型式",
            "自動車登録番号",
            "新規登録年月",
            "自動車種類",
            "車体の形状",
            "車名",
            "最大積載量(kg)",
            "リース有無",
            "リース金額\n(月額)",
            "リース\n終了年月",
            "減価償却費",
            "リース種類",
            "リース金額\n(月額)",
            "リース残額\n(月額)",
            "取得価額",
            "簿価額",
            "備考1",
            "備考2",
            "備考3",
        ]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)

        return wb

    def _write_to_excel(self, wb: Workbook, vehicles_data: List[Dict]):
        """
        車両データをExcelに書き込み

        Args:
            wb: Workbookオブジェクト
            vehicles_data: 車両情報のリスト
        """
        # 車両台帳シートを取得
        ws = wb["車両台帳"]

        # データは3行目から書き込み
        start_row = 3

        # 列マッピング(A列=1)
        column_map = {
            "No": 1,
            "車両名": 3,
            "車台番号": 4,
            "型式": 5,
            "自動車登録番号": 6,
            "新規登録年月": 7,
            "自動車種類": 8,
            "車体の形状": 9,
            "車名": 10,
            "最大積載量(kg)": 11,
        }

        # データを書き込み
        for idx, vehicle in enumerate(vehicles_data):
            row = start_row + idx
            for field, col in column_map.items():
                value = vehicle.get(field, "")
                ws.cell(row=row, column=col, value=value)
