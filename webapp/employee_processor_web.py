"""
従業員台帳プロセッサ（Web版）
アップロードされたファイルから従業員情報を抽出してExcelに書き込む
"""
from io import BytesIO
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from utils.claude_client import ClaudeClient


class EmployeeProcessorWeb:
    """従業員台帳プロセッサ（Web版）"""

    def __init__(self):
        self.claude_client = ClaudeClient()

    def process_uploaded_files(
        self, uploaded_files: List, company_name: str
    ) -> BytesIO:
        """
        アップロードされたファイルから従業員台帳Excelを生成

        Args:
            uploaded_files: Streamlit UploadedFileのリスト
            company_name: 会社名

        Returns:
            BytesIO: Excelファイルのバイトストリーム
        """
        employees_data = []
        employee_no = 1

        # 各ファイルを処理
        for uploaded_file in uploaded_files:
            file_bytes = uploaded_file.read()
            filename = uploaded_file.name

            # 従業員情報抽出
            if filename.lower().endswith('.pdf'):
                employee_info = self.claude_client.analyze_pdf_bytes(
                    file_bytes, self._get_employee_prompt()
                )
            elif filename.lower().endswith(('.xlsx', '.xls')):
                # Excelファイルの場合もPDFとして処理（画像化）
                employee_info = self.claude_client.analyze_pdf_bytes(
                    file_bytes, self._get_employee_prompt()
                )
            else:
                # 画像ファイル
                employee_info = self.claude_client.analyze_image_bytes(
                    file_bytes, filename, self._get_employee_prompt()
                )

            # データ整形
            if isinstance(employee_info, list):
                for info in employee_info:
                    info["No"] = employee_no
                    employees_data.append(info)
                    employee_no += 1
            elif isinstance(employee_info, dict) and not employee_info.get("error"):
                employee_info["No"] = employee_no
                employees_data.append(employee_info)
                employee_no += 1

        # Excelワークブック作成
        wb = self._create_workbook(company_name)
        self._write_to_excel(wb, employees_data)

        # BytesIOに保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def _get_employee_prompt(self) -> str:
        """従業員情報解析用プロンプト"""
        return """
この資料から従業員情報を抽出してJSONで返してください。
情報が読み取れない場合は空文字列""を設定してください。
複数名の従業員情報がある場合は、すべて抽出して配列で返してください。

抽出する情報:
- 氏名（フルネーム）
- 年齢（数値のみ、例: 35）
- 性別（男性/女性）
- 役職（例: 代表取締役、部長、一般社員、パート）
- 入社日（例: 2020年4月1日、令和2年4月）
- 給与_月額（数値のみ、例: 350000）
- 賞与（年額、数値のみ、例: 700000）
- 雇用形態（正社員/契約社員/パート/アルバイト）
- 勤務地（例: 本社、東京支店）
- 備考（その他特記事項）

JSONフォーマット(1名の場合):
{
  "氏名": "山田太郎",
  "年齢": 35,
  "性別": "男性",
  "役職": "部長",
  "入社日": "2015年4月1日",
  "給与_月額": 450000,
  "賞与": 900000,
  "雇用形態": "正社員",
  "勤務地": "本社",
  "備考": ""
}

複数名の場合は配列で返してください:
[
  {...},
  {...}
]

重要:
- 数値は数字のみ（カンマや円マークは不要）
- 年齢、給与、賞与は数値型で返す
- 情報がない項目は空文字列""または0を設定
- 必ずJSON形式で返してください（コードブロック不要）
"""

    def _create_workbook(self, company_name: str) -> Workbook:
        """
        新規Workbookを作成

        Args:
            company_name: 会社名

        Returns:
            Workbookオブジェクト
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "従業員台帳"

        # タイトル行（1行目）
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"{company_name} 従業員台帳"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # ヘッダー行（2行目）
        headers = [
            "No",
            "氏名",
            "年齢",
            "性別",
            "役職",
            "入社日",
            "給与（月額）",
            "賞与（年額）",
            "雇用形態",
            "勤務地",
            "備考",
        ]

        # ヘッダースタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 列幅調整
        column_widths = {
            'A': 6,   # No
            'B': 15,  # 氏名
            'C': 8,   # 年齢
            'D': 8,   # 性別
            'E': 15,  # 役職
            'F': 15,  # 入社日
            'G': 12,  # 給与
            'H': 12,  # 賞与
            'I': 12,  # 雇用形態
            'J': 15,  # 勤務地
            'K': 20,  # 備考
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 2行目の高さ調整
        ws.row_dimensions[2].height = 30

        return wb

    def _write_to_excel(self, wb: Workbook, employees_data: List[Dict]):
        """
        従業員データをExcelに書き込み

        Args:
            wb: Workbookオブジェクト
            employees_data: 従業員情報のリスト
        """
        ws = wb["従業員台帳"]

        # データは3行目から書き込み
        start_row = 3

        # 列マッピング(A列=1)
        column_map = {
            "No": 1,
            "氏名": 2,
            "年齢": 3,
            "性別": 4,
            "役職": 5,
            "入社日": 6,
            "給与_月額": 7,
            "賞与": 8,
            "雇用形態": 9,
            "勤務地": 10,
            "備考": 11,
        }

        # 罫線スタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # データを書き込み
        for idx, employee in enumerate(employees_data):
            row = start_row + idx

            for field, col in column_map.items():
                cell = ws.cell(row=row, column=col)
                value = employee.get(field, "")

                # 数値フィールドの処理
                if field in ["年齢", "給与_月額", "賞与"]:
                    try:
                        cell.value = int(value) if value else 0
                    except (ValueError, TypeError):
                        cell.value = 0
                else:
                    cell.value = value

                # セルスタイル
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center' if field in ["No", "年齢", "性別", "雇用形態"] else 'left',
                    vertical='center'
                )

        # 合計行を追加
        total_row = start_row + len(employees_data)
        ws.cell(total_row, 1).value = "合計"
        ws.cell(total_row, 1).font = Font(bold=True)
        ws.cell(total_row, 3).value = len(employees_data)
        ws.cell(total_row, 3).font = Font(bold=True)
        ws.cell(total_row, 3).alignment = Alignment(horizontal='center')
