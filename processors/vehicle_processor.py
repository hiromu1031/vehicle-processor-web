"""
車両台帳プロセッサ
車検証画像から車両情報を抽出してExcelに書き込む
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl import load_workbook

from ..utils.claude_client import ClaudeClient
from ..utils.excel_handler import ExcelHandler
from ..config import VEHICLE_SCAN_FOLDER, VEHICLE_OUTPUT_FOLDER, VEHICLE_EXCEL_NAME


class VehicleProcessor:
    """車両台帳プロセッサ"""

    def __init__(self):
        self.claude_client = ClaudeClient()
        self.excel_handler = ExcelHandler()

    def process_company(self, company_folder: str) -> str:
        """
        企業フォルダから車検証を処理して車両台帳を生成

        Args:
            company_folder: 企業フォルダパス

        Returns:
            生成されたExcelファイルパス
        """
        company_path = Path(company_folder)
        company_name = self._extract_company_name(company_path.name)

        print(f"\n{'='*60}")
        print(f"車両台帳処理開始: {company_name}")
        print(f"{'='*60}\n")

        # スキャンフォルダから車検証を取得 (複数のフォルダ構造に対応)
        scan_candidates = [
            company_path / "スキャン（その他）" / VEHICLE_SCAN_FOLDER,
            company_path / "スキャン" / VEHICLE_SCAN_FOLDER,
            company_path / "0_共有フォルダ" / VEHICLE_SCAN_FOLDER,
        ]

        scan_folder = None
        for candidate in scan_candidates:
            if candidate.exists():
                scan_folder = candidate
                break

        if not scan_folder:
            print(f"警告: 車検証フォルダが見つかりません")
            return None

        vehicle_files = (
            list(scan_folder.glob("*.jpg"))
            + list(scan_folder.glob("*.jpeg"))
            + list(scan_folder.glob("*.png"))
            + list(scan_folder.glob("*.pdf"))
        )

        if not vehicle_files:
            print(f"警告: 車検証ファイルが見つかりません")
            return None

        print(f"[OK] 車検証ファイル数: {len(vehicle_files)}")

        # 各車検証から情報を抽出
        vehicles_data = []
        vehicle_no = 1
        for idx, vehicle_file in enumerate(vehicle_files, start=1):
            print(f"\n[{idx}/{len(vehicle_files)}] 処理中: {vehicle_file.name}")
            try:
                vehicle_info = self._extract_vehicle_info(str(vehicle_file))
                if not vehicle_info:
                    print(f"  [NG] 抽出失敗: データなし")
                    continue

                # 複数台の場合(配列)と単一台の場合(辞書)を処理
                if isinstance(vehicle_info, list):
                    if len(vehicle_info) > 0:
                        for info in vehicle_info:
                            info["No"] = vehicle_no
                            vehicles_data.append(info)
                            print(f"  [OK] 抽出成功: {info.get('車両名', '不明')}")
                            vehicle_no += 1
                    else:
                        print(f"  [NG] 抽出失敗: 空の配列")
                elif isinstance(vehicle_info, dict):
                    if vehicle_info.get("error"):
                        print(f"  [NG] 抽出失敗: {vehicle_info.get('error')}")
                    else:
                        vehicle_info["No"] = vehicle_no
                        vehicles_data.append(vehicle_info)
                        print(f"  [OK] 抽出成功: {vehicle_info.get('車両名', '不明')}")
                        vehicle_no += 1
                else:
                    print(f"  [NG] 抽出失敗: 予期しない形式")
            except Exception as e:
                print(f"  [NG] エラー: {e}")

        if not vehicles_data:
            print("\n警告: 抽出できた車両情報がありません")
            return None

        # テンプレートExcelを読み込み
        template_path = self._find_template(company_path)
        if template_path:
            print(f"\n[OK] テンプレート使用: {template_path}")
            wb = load_workbook(template_path)
        else:
            print(f"\n新規作成します")
            wb = self._create_new_workbook(company_name)

        # データを書き込み
        self._write_to_excel(wb, vehicles_data)

        # 保存
        output_folder = company_path / VEHICLE_OUTPUT_FOLDER
        output_folder.mkdir(parents=True, exist_ok=True)
        output_path = output_folder / VEHICLE_EXCEL_NAME.format(
            company_name=company_name
        )

        self.excel_handler.save_workbook(wb, str(output_path))
        print(f"\n[OK] 車両台帳生成完了: {len(vehicles_data)}台")

        return str(output_path)

    def _extract_vehicle_info(self, file_path: str) -> Dict[str, Any]:
        """
        車検証ファイル(画像またはPDF)から車両情報を抽出

        Args:
            file_path: 車検証ファイルパス

        Returns:
            車両情報(辞書)
        """
        prompt = """
この車検証から以下の情報を抽出してJSONで返してください。
情報が読み取れない場合は空文字列""を設定してください。
複数台の車両情報がある場合は、すべて抽出して配列で返してください。

抽出する情報:
- 車両名 (例: ダイハツ ハイゼット)
- 車台番号
- 型式
- 自動車登録番号 (ナンバープレート)
- 新規登録年月
- 自動車種類 (例: 貨物、乗用)
- 車体の形状 (例: バン)
- 車名
- 最大積載量(kg)

JSONフォーマット(1台の場合):
{
  "車両名": "ダイハツ ハイゼット",
  "車台番号": "S331V-0046341",
  "型式": "EBD-S331V",
  "自動車登録番号": "群馬 480 ま 8601",
  "新規登録年月": "令和23年4月",
  "自動車種類": "貨物",
  "車体の形状": "バン",
  "車名": "ダイハツ",
  "最大積載量(kg)": "350(250)"
}

複数台の場合は配列で返してください:
[
  {...},
  {...}
]

必ずJSON形式で返してください。コードブロックは使わず、JSONのみを返してください。
"""

        # ファイル拡張子で判定
        if file_path.lower().endswith('.pdf'):
            return self.claude_client.analyze_pdf(
                file_path, prompt, response_format="json"
            )
        else:
            return self.claude_client.analyze_image(
                file_path, prompt, response_format="json"
            )

    def _write_to_excel(self, wb: openpyxl.Workbook, vehicles_data: List[Dict]):
        """
        車両データをExcelに書き込み

        Args:
            wb: Workbookオブジェクト
            vehicles_data: 車両情報のリスト
        """
        # 車両台帳シートを取得または作成
        if "車両台帳" in wb.sheetnames:
            ws = wb["車両台帳"]
        else:
            ws = wb.create_sheet("車両台帳")

        # ヘッダー行を確認(2行目にあると仮定)
        # データは3行目から書き込み
        start_row = 3

        # 列マッピング(A列=1)
        column_map = {
            "No": 1,
            "車両名": 3,
            "車台番号": 4,
            "型式": 5,
            "自動車登録番号": 6,
            "新規登録年月": 7,
            "自動車種類": 8,
            "車体の形状": 9,
            "車名": 10,
            "最大積載量(kg)": 11,
        }

        # データを書き込み
        for idx, vehicle in enumerate(vehicles_data):
            row = start_row + idx
            for field, col in column_map.items():
                value = vehicle.get(field, "")
                ws.cell(row=row, column=col, value=value)

    def _find_template(self, company_path: Path) -> str:
        """
        既存の車両台帳テンプレートを探す

        Args:
            company_path: 企業フォルダパス

        Returns:
            テンプレートパスまたはNone
        """
        output_folder = company_path / VEHICLE_OUTPUT_FOLDER
        if output_folder.exists():
            excel_files = list(output_folder.glob("*.xlsx"))
            if excel_files:
                return str(excel_files[0])
        return None

    def _create_new_workbook(self, company_name: str) -> openpyxl.Workbook:
        """
        新規Workbookを作成

        Args:
            company_name: 会社名

        Returns:
            Workbookオブジェクト
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "車両台帳"

        # ヘッダー行を作成(2行目)
        headers = [
            "No",
            "ステータス",
            "車両名",
            "車台番号",
            "型式",
            "自動車登録番号",
            "新規登録年月",
            "自動車種類",
            "車体の形状",
            "車名",
            "最大積載量(kg)",
            "リース有無",
            "リース金額\n(月額)",
            "リース\n終了年月",
            "減価償却費",
            "リース種類",
            "リース金額\n(月額)",
            "リース残額\n(月額)",
            "取得価額",
            "簿価額",
            "備考1",
            "備考2",
            "備考3",
        ]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)

        return wb

    def _extract_company_name(self, folder_name: str) -> str:
        """
        フォルダ名から会社名を抽出

        Args:
            folder_name: フォルダ名 (例: "A918_株式会社野木工業")

        Returns:
            会社名
        """
        if "_" in folder_name:
            return folder_name.split("_", 1)[1]
        return folder_name
