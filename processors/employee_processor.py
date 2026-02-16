"""
従業員台帳プロセッサ
従業員名簿・履歴書画像から従業員情報を抽出してExcelに書き込む
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl import load_workbook

from ..utils.claude_client import ClaudeClient
from ..utils.excel_handler import ExcelHandler
from ..config import EMPLOYEE_SCAN_FOLDER, EMPLOYEE_OUTPUT_FOLDER, EMPLOYEE_EXCEL_NAME


class EmployeeProcessor:
    """従業員台帳プロセッサ"""

    def __init__(self):
        self.claude_client = ClaudeClient()
        self.excel_handler = ExcelHandler()

    def process_company(self, company_folder: str) -> str:
        """
        企業フォルダから従業員名簿を処理して従業員台帳を生成

        Args:
            company_folder: 企業フォルダパス

        Returns:
            生成されたExcelファイルパス
        """
        company_path = Path(company_folder)
        company_name = self._extract_company_name(company_path.name)

        print(f"\n{'='*60}")
        print(f"従業員台帳処理開始: {company_name}")
        print(f"{'='*60}\n")

        # スキャンフォルダから従業員名簿を取得 (複数のフォルダ構造に対応)
        scan_candidates = [
            company_path / "スキャン（その他）" / EMPLOYEE_SCAN_FOLDER,
            company_path / "スキャン" / EMPLOYEE_SCAN_FOLDER,
            company_path / "0_共有フォルダ" / EMPLOYEE_SCAN_FOLDER,
            company_path / "スキャン（その他）" / (EMPLOYEE_SCAN_FOLDER + "〇"),
            company_path / "スキャン" / (EMPLOYEE_SCAN_FOLDER + "〇"),
            company_path / "0_共有フォルダ" / (EMPLOYEE_SCAN_FOLDER + "〇"),
        ]

        scan_folder = None
        for candidate in scan_candidates:
            if candidate.exists():
                scan_folder = candidate
                break

        if not scan_folder:
            print(f"警告: 従業員名簿フォルダが見つかりません")
            return None

        employee_files = (
            list(scan_folder.glob("*.jpg"))
            + list(scan_folder.glob("*.jpeg"))
            + list(scan_folder.glob("*.png"))
            + list(scan_folder.glob("*.pdf"))
        )

        if not employee_files:
            print(f"警告: 従業員名簿ファイルが見つかりません")
            return None

        print(f"[OK] 従業員名簿ファイル数: {len(employee_files)}")

        # ファイルを画像とPDFに分類
        image_files = [
            f
            for f in employee_files
            if f.suffix.lower() in [".jpg", ".jpeg", ".png"]
        ]
        pdf_files = [f for f in employee_files if f.suffix.lower() == ".pdf"]

        employees_data = []

        # 画像ファイルを処理
        if image_files:
            print(f"\n画像ファイル {len(image_files)}枚を一括解析中...")
            try:
                data = self._extract_employees_info([str(f) for f in image_files])
                if data and isinstance(data, list) and len(data) > 0:
                    employees_data.extend(data)
                    print(f"[OK] 抽出成功: {len(data)}名")
                else:
                    print(f"[NG] 抽出失敗")
            except Exception as e:
                print(f"[NG] エラー: {e}")

        # PDFファイルを処理
        if pdf_files:
            print(f"\nPDFファイル {len(pdf_files)}個を解析中...")
            for pdf_file in pdf_files:
                try:
                    print(f"  処理中: {pdf_file.name}")
                    data = self._extract_employees_from_pdf(str(pdf_file))
                    if data and isinstance(data, list) and len(data) > 0:
                        employees_data.extend(data)
                        print(f"  [OK] 抽出成功: {len(data)}名")
                    else:
                        print(f"  [NG] 抽出失敗")
                except Exception as e:
                    print(f"  [NG] エラー: {e}")

        if not employees_data:
            print("\n警告: 抽出できた従業員情報がありません")
            return None

        # Noを付与
        for idx, emp in enumerate(employees_data, start=1):
            emp["No"] = idx

        # テンプレートExcelを読み込み
        template_path = self._find_template(company_path)
        if template_path:
            print(f"\n[OK] テンプレート使用: {template_path}")
            wb = load_workbook(template_path)
        else:
            print(f"\n新規作成します")
            wb = self._create_new_workbook(company_name)

        # データを書き込み
        self._write_to_excel(wb, employees_data)

        # 保存
        output_folder = company_path / EMPLOYEE_OUTPUT_FOLDER
        output_folder.mkdir(parents=True, exist_ok=True)
        output_path = output_folder / EMPLOYEE_EXCEL_NAME.format(
            company_name=company_name
        )

        self.excel_handler.save_workbook(wb, str(output_path))
        print(f"\n[OK] 従業員台帳生成完了: {len(employees_data)}名")

        return str(output_path)

    def _extract_employees_info(self, image_paths: List[str]) -> List[Dict[str, Any]]:
        """
        従業員名簿画像から従業員情報を抽出

        Args:
            image_paths: 従業員名簿画像パスのリスト

        Returns:
            従業員情報のリスト
        """
        prompt = """
これらの従業員関連書類(名簿・履歴書・合格証など)から従業員情報を抽出してJSONで返してください。

抽出する情報(各従業員について):
- 氏名
- 生年月日 (YYYY年MM月DD日形式、または元号形式)
- 年齢 (計算できる場合)
- 雇用形態 (正社員、パート、契約社員など。記載がない場合は空文字列)
- 保有資格 (資格名をカンマ区切り。記載がない場合は空文字列)
- 職種 (代表取締役、技術者など。記載がない場合は空文字列)
- 備考 (その他の情報。記載がない場合は空文字列)

JSONフォーマット:
[
  {
    "氏名": "野木 和彦",
    "生年月日": "昭和53年1月25日",
    "年齢": 47,
    "雇用形態": "",
    "保有資格": "一級土木施工管理技士、一級建築施工管理技士",
    "職種": "代表取締役",
    "備考": ""
  },
  {
    "氏名": "山田 太郎",
    "生年月日": "1990年5月10日",
    "年齢": 35,
    "雇用形態": "正社員",
    "保有資格": "",
    "職種": "",
    "備考": ""
  }
]

必ずJSON配列形式で返してください。コードブロックは使わず、JSONのみを返してください。
複数の従業員が見つかった場合は、すべてを配列に含めてください。
"""

        result = self.claude_client.analyze_multiple_images(
            image_paths, prompt, response_format="json"
        )

        # resultがリストで返ってくることを期待
        if isinstance(result, list):
            return result
        elif isinstance(result, dict) and "error" in result:
            return []
        else:
            # 予期しないフォーマットの場合
            print(f"警告: 予期しないレスポンス形式: {type(result)}")
            return []

    def _extract_employees_from_pdf(self, pdf_path: str) -> List[Dict[str, Any]]:
        """
        従業員名簿PDFから従業員情報を抽出

        Args:
            pdf_path: 従業員名簿PDFパス

        Returns:
            従業員情報のリスト
        """
        prompt = """
この従業員名簿PDFから従業員情報を抽出してJSONで返してください。

抽出する情報(各従業員について):
- 氏名
- 生年月日 (YYYY年MM月DD日形式、または元号形式)
- 年齢 (計算できる場合)
- 雇用形態 (正社員、パート、契約社員など。記載がない場合は空文字列)
- 保有資格 (資格名をカンマ区切り。記載がない場合は空文字列)
- 職種 (代表取締役、技術者など。記載がない場合は空文字列)
- 備考 (その他の情報。記載がない場合は空文字列)

JSONフォーマット:
[
  {
    "氏名": "野木 和彦",
    "生年月日": "昭和53年1月25日",
    "年齢": 47,
    "雇用形態": "正社員",
    "保有資格": "",
    "職種": "運転手",
    "備考": ""
  }
]

必ずJSON配列形式で返してください。コードブロックは使わず、JSONのみを返してください。
すべての従業員を配列に含めてください。
"""

        result = self.claude_client.analyze_pdf(
            pdf_path, prompt, response_format="json"
        )

        if isinstance(result, list):
            return result
        elif isinstance(result, dict) and "error" in result:
            return []
        else:
            print(f"警告: 予期しないレスポンス形式: {type(result)}")
            return []

    def _write_to_excel(self, wb: openpyxl.Workbook, employees_data: List[Dict]):
        """
        従業員データをExcelに書き込み

        Args:
            wb: Workbookオブジェクト
            employees_data: 従業員情報のリスト
        """
        # 従業員一覧シートを取得または作成
        if "従業員一覧" in wb.sheetnames:
            ws = wb["従業員一覧"]
        else:
            ws = wb.create_sheet("従業員一覧")

        # データは2行目から書き込み(1行目はヘッダー)
        start_row = 2

        # 列マッピング
        column_map = {
            "No": 1,
            "氏名": 2,
            "生年月日": 3,
            "年齢": 4,
            "雇用形態": 5,
            "保有資格": 6,
            "職種": 7,
            "備考": 8,
        }

        # データを書き込み
        for idx, employee in enumerate(employees_data):
            row = start_row + idx
            for field, col in column_map.items():
                value = employee.get(field, "")
                ws.cell(row=row, column=col, value=value)

    def _find_template(self, company_path: Path) -> str:
        """
        既存の従業員台帳テンプレートを探す

        Args:
            company_path: 企業フォルダパス

        Returns:
            テンプレートパスまたはNone
        """
        output_folder = company_path / EMPLOYEE_OUTPUT_FOLDER
        if output_folder.exists():
            excel_files = list(output_folder.glob("*.xlsx"))
            if excel_files:
                return str(excel_files[0])
        return None

    def _create_new_workbook(self, company_name: str) -> openpyxl.Workbook:
        """
        新規Workbookを作成

        Args:
            company_name: 会社名

        Returns:
            Workbookオブジェクト
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "従業員一覧"

        # ヘッダー行を作成
        headers = ["No", "氏名", "生年月日", "年齢", "雇用形態", "保有資格", "職種", "備考"]

        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        return wb

    def _extract_company_name(self, folder_name: str) -> str:
        """
        フォルダ名から会社名を抽出

        Args:
            folder_name: フォルダ名

        Returns:
            会社名
        """
        if "_" in folder_name:
            return folder_name.split("_", 1)[1]
        return folder_name
