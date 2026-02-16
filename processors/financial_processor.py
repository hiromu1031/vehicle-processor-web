"""
財務概要プロセッサ
決算書PDFから財務情報を抽出してExcelに書き込む
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl import load_workbook

from ..utils.claude_client import ClaudeClient
from ..utils.excel_handler import ExcelHandler
from ..config import FINANCIAL_SCAN_FOLDER, FINANCIAL_OUTPUT_FOLDER, FINANCIAL_EXCEL_NAME


class FinancialProcessor:
    """財務概要プロセッサ"""

    def __init__(self):
        self.claude_client = ClaudeClient()
        self.excel_handler = ExcelHandler()

    def process_company(self, company_folder: str) -> str:
        """
        企業フォルダから決算書を処理して財務概要を生成

        Args:
            company_folder: 企業フォルダパス

        Returns:
            生成されたExcelファイルパス
        """
        company_path = Path(company_folder)
        company_name = self._extract_company_name(company_path.name)

        print(f"\n{'='*60}")
        print(f"財務概要処理開始: {company_name}")
        print(f"{'='*60}\n")

        # スキャンフォルダから決算書を取得 (複数のフォルダ構造に対応)
        scan_candidates = [
            company_path / "スキャン（その他）" / FINANCIAL_SCAN_FOLDER,
            company_path / "スキャン" / FINANCIAL_SCAN_FOLDER,
            company_path / "スキャン（その他）",
            company_path / "スキャン",
        ]

        scan_folder = None
        for candidate in scan_candidates:
            if candidate.exists():
                # 決算書PDFがあるか確認
                pdf_files = list(candidate.glob("*.pdf"))
                if pdf_files:
                    scan_folder = candidate
                    break

        if not scan_folder:
            print(f"警告: 決算書フォルダまたは決算書PDFが見つかりません")
            return None

        pdf_files = list(scan_folder.glob("*.pdf"))

        if not pdf_files:
            print(f"警告: 決算書PDFファイルが見つかりません")
            return None

        print(f"[OK] 決算書PDFファイル数: {len(pdf_files)}")

        # 各決算書から情報を抽出
        print(f"\n決算書PDFを一括解析中...")
        try:
            financial_data = self._extract_financial_info_batch(
                [str(f) for f in pdf_files]
            )
            if financial_data and not financial_data.get("error"):
                print(f"[OK] 抽出成功")
            else:
                print(f"[NG] 抽出失敗")
                return None
        except Exception as e:
            print(f"[NG] エラー: {e}")
            return None

        # テンプレートExcelを読み込み
        template_path = self._find_template(company_path)
        if template_path:
            print(f"\n[OK] テンプレート使用: {template_path}")
            wb = load_workbook(template_path)
        else:
            print(f"\n新規作成します")
            wb = self._create_new_workbook(company_name)

        # データを書き込み
        self._write_to_excel(wb, financial_data)

        # 保存
        output_folder = company_path / FINANCIAL_OUTPUT_FOLDER
        output_folder.mkdir(parents=True, exist_ok=True)
        output_path = output_folder / FINANCIAL_EXCEL_NAME.format(
            company_name=company_name
        )

        self.excel_handler.save_workbook(wb, str(output_path))
        print(f"\n[OK] 財務概要生成完了")

        return str(output_path)

    def _extract_financial_info_batch(
        self, pdf_paths: List[str]
    ) -> Dict[str, Any]:
        """
        決算書PDFから財務情報を一括抽出

        Args:
            pdf_paths: 決算書PDFパスのリスト

        Returns:
            財務情報(PL, BS, 科目データ)
        """
        # 各PDFをチェックして、PL/BSを含むものを探す
        check_prompt = "このPDFに損益計算書(PL)または貸借対照表(BS)は含まれていますか? 含まれている場合は「はい」、含まれていない場合は「いいえ」と答えてください。"

        pl_bs_pdfs = []
        print(f"\n各PDFの内容をチェック中...")

        for idx, pdf_path in enumerate(pdf_paths[:10], start=1):  # 最初の10個まで
            try:
                print(f"  [{idx}/{min(10, len(pdf_paths))}] {Path(pdf_path).name} をチェック中...")
                result = self.claude_client.analyze_pdf(
                    pdf_path, check_prompt, response_format="text"
                )
                response = result.get("text", "").lower()

                if "はい" in response or "yes" in response or "含まれ" in response:
                    print(f"    → PL/BS発見!")
                    pl_bs_pdfs.append(pdf_path)
                else:
                    print(f"    → PL/BSなし")
            except Exception as e:
                print(f"    → エラー: {e}")
                continue

        if not pl_bs_pdfs:
            print(f"\n警告: PL/BSを含むPDFが見つかりませんでした")
            # 最初のPDFで試す
            pl_bs_pdfs = [pdf_paths[0]]

        print(f"\n財務情報を抽出中...")

        # PL/BSを含むPDFから情報を抽出
        all_financial_data = {
            "年度": [],
            "PL": {},
            "BS": {}
        }

        for pdf_path in pl_bs_pdfs[:3]:  # 最大3つのPDFを処理
            print(f"  処理中: {Path(pdf_path).name}")

            prompt = """
この決算書PDFから以下の財務情報を抽出してJSONで返してください。

抽出する情報:
1. 損益計算書(PL): 売上高、売上原価、売上総利益、販管費、営業利益、経常利益、当期純利益など
2. 貸借対照表(BS): 流動資産、固定資産、資産合計、流動負債、固定負債、負債合計、純資産など
3. 対象年度: 第X期、決算年月日

複数年度のデータがある場合は、すべての年度を抽出してください。

JSONフォーマット:
{
  "年度": [
    {"期": "第2期", "決算年月日": "2022年8月31日"},
    {"期": "第3期", "決算年月日": "2023年8月31日"}
  ],
  "PL": {
    "売上高": [52490777, 45916331],
    "売上原価": [35349283, 25568449],
    "売上総利益": [17141494, 20347882],
    "販管費": [14115022, 13179901],
    "営業利益": [3026472, 7167981],
    "経常利益": [2965632, 7105481],
    "当期純利益": [1931700, 4636600]
  },
  "BS": {
    "流動資産": [23145445, 23042785],
    "固定資産": [5824321, 6234567],
    "資産合計": [28969766, 29277352],
    "流動負債": [12345678, 11234567],
    "固定負債": [5678901, 6789012],
    "負債合計": [18024579, 18023579],
    "純資産": [10945187, 11253773]
  }
}

必ずJSON形式で返してください。
数値はカンマなしの数値として返してください。
損益計算書や貸借対照表が含まれていない場合は、空の配列を返してください。
"""

            try:
                data = self.claude_client.analyze_pdf(
                    pdf_path, prompt, response_format="json"
                )

                # データをマージ
                if data.get("年度"):
                    all_financial_data["年度"].extend(data["年度"])

                if data.get("PL"):
                    for key, values in data["PL"].items():
                        if key not in all_financial_data["PL"]:
                            all_financial_data["PL"][key] = []
                        all_financial_data["PL"][key].extend(values)

                if data.get("BS"):
                    for key, values in data["BS"].items():
                        if key not in all_financial_data["BS"]:
                            all_financial_data["BS"][key] = []
                        all_financial_data["BS"][key].extend(values)

                print(f"    → 抽出成功")
            except Exception as e:
                print(f"    → エラー: {e}")
                continue

        return all_financial_data

    def _write_to_excel(self, wb: openpyxl.Workbook, financial_data: Dict):
        """
        財務データをExcelに書き込み

        Args:
            wb: Workbookオブジェクト
            financial_data: 財務情報
        """
        # PLシートに書き込み
        if "PL" in financial_data:
            self._write_pl_sheet(wb, financial_data)

        # BSシートに書き込み
        if "BS" in financial_data:
            self._write_bs_sheet(wb, financial_data)

    def _write_pl_sheet(self, wb: openpyxl.Workbook, financial_data: Dict):
        """PLシートにデータを書き込み"""
        if "PL" not in wb.sheetnames:
            ws = wb.create_sheet("PL")
        else:
            ws = wb["PL"]

        pl_data = financial_data.get("PL", {})
        年度 = financial_data.get("年度", [])

        # ヘッダー行を作成(1行目)
        ws.cell(row=1, column=3, value="勘定科目")
        for idx, 期 in enumerate(年度, start=1):
            col = 2 + idx
            ws.cell(row=1, column=col, value=期.get("決算年月日", ""))

        # PL科目のマッピング
        pl_items = [
            ("売上高", "売上高"),
            ("売上原価", "売上原価"),
            ("売上総利益", "売上総利益"),
            ("販管費", "販管費"),
            ("営業利益", "営業利益"),
            ("経常利益", "経常利益"),
            ("当期純利益", "当期純利益"),
        ]

        row = 2
        for label, key in pl_items:
            ws.cell(row=row, column=2, value=label)
            values = pl_data.get(key, [])
            for idx, value in enumerate(values):
                col = 3 + idx
                ws.cell(row=row, column=col, value=value)
            row += 1

    def _write_bs_sheet(self, wb: openpyxl.Workbook, financial_data: Dict):
        """BSシートにデータを書き込み"""
        if "BS" not in wb.sheetnames:
            ws = wb.create_sheet("BS")
        else:
            ws = wb["BS"]

        bs_data = financial_data.get("BS", {})
        年度 = financial_data.get("年度", [])

        # ヘッダー行を作成
        ws.cell(row=1, column=3, value="勘定科目")
        for idx, 期 in enumerate(年度, start=1):
            col = 2 + idx
            ws.cell(row=1, column=col, value=期.get("決算年月日", ""))

        # BS科目のマッピング
        bs_items = [
            ("流動資産", "流動資産"),
            ("固定資産", "固定資産"),
            ("資産合計", "資産合計"),
            ("流動負債", "流動負債"),
            ("固定負債", "固定負債"),
            ("負債合計", "負債合計"),
            ("純資産", "純資産"),
        ]

        row = 2
        for label, key in bs_items:
            ws.cell(row=row, column=2, value=label)
            values = bs_data.get(key, [])
            for idx, value in enumerate(values):
                col = 3 + idx
                ws.cell(row=row, column=col, value=value)
            row += 1

    def _find_template(self, company_path: Path) -> str:
        """既存の財務概要テンプレートを探す"""
        output_folder = company_path / FINANCIAL_OUTPUT_FOLDER
        if output_folder.exists():
            excel_files = list(output_folder.glob("*.xlsx"))
            if excel_files:
                # テンプレートという名前のファイルを優先
                for f in excel_files:
                    if "テンプレート" in f.name:
                        return str(f)
                # なければ最初のファイル
                return str(excel_files[0])
        return None

    def _create_new_workbook(self, company_name: str) -> openpyxl.Workbook:
        """新規Workbookを作成"""
        wb = openpyxl.Workbook()

        # PLシート
        ws_pl = wb.active
        ws_pl.title = "PL"

        # BSシート
        ws_bs = wb.create_sheet("BS")

        # 科目シート
        ws_accounts = wb.create_sheet("科目")

        return wb

    def _extract_company_name(self, folder_name: str) -> str:
        """フォルダ名から会社名を抽出"""
        if "_" in folder_name:
            return folder_name.split("_", 1)[1]
        return folder_name
