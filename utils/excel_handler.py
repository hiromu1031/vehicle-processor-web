"""
Excel操作ユーティリティ
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


class ExcelHandler:
    """Excelファイル操作クラス"""

    @staticmethod
    def load_template(template_path: str) -> Workbook:
        """
        テンプレートExcelを読み込み

        Args:
            template_path: テンプレートファイルパス

        Returns:
            Workbookオブジェクト
        """
        path = Path(template_path)
        if path.exists():
            return load_workbook(template_path)
        else:
            # テンプレートが存在しない場合は新規作成
            return Workbook()

    @staticmethod
    def write_data_to_sheet(
        wb: Workbook,
        sheet_name: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        start_row: int = 2,
    ):
        """
        データをシートに書き込み

        Args:
            wb: Workbookオブジェクト
            sheet_name: シート名
            data: 書き込むデータ(辞書のリスト)
            headers: ヘッダー列名のリスト
            start_row: データ書き込み開始行(1-indexed)
        """
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # データを書き込み
        for row_idx, item in enumerate(data, start=start_row):
            for col_idx, header in enumerate(headers, start=1):
                value = item.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 基本的なスタイル設定
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

    @staticmethod
    def save_workbook(wb: Workbook, output_path: str):
        """
        Workbookを保存

        Args:
            wb: Workbookオブジェクト
            output_path: 出力ファイルパス
        """
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"[OK] Excelファイルを保存しました: {output_path}")
