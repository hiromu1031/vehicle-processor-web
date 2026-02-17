"""
財務概要Excelファイルの構造を分析
"""
import openpyxl
from pathlib import Path

excel_path = r"C:\Users\ひろむpcマスタ\Desktop\claude\A915_久栄運輸株式会社\決算書打ち込み\財務概要_久栄運輸株式会社.xlsx"

wb = openpyxl.load_workbook(excel_path)

print("=" * 80)
print("シート一覧:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

print("\n" + "=" * 80)

# 各シートの内容を詳細に確認
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n【{sheet_name}】シート:")
    print(f"  最大行: {ws.max_row}, 最大列: {ws.max_column}")

    # 最初の20行を表示
    print("\n  --- 最初の20行 ---")
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(15, ws.max_column + 1)):  # 最初の15列まで
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                row_data.append(f"[{col_idx}]{value}")
        if row_data:
            print(f"  行{row_idx}: {' | '.join(row_data)}")

    print("\n" + "-" * 80)

wb.close()
print("\n分析完了")
